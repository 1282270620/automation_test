from openpyxl import load_workbook
from openpyxl import Workbook
# 如何读取Excel
# workbook = load_workbook("c:\\test.xlsx")
# worksheet = workbook.get_sheet_by_name("Sheet1")
# # 得到一列的单元格对象
# print(worksheet["A"])
# # 得到一行的单元格对象
# print(worksheet["1"])
# # 得到一行单元格中第一个的值
# print(worksheet["1"][0].value)
# # 得到最大行
# print(worksheet.max_row)
# # 得到最大列
# print(worksheet.max_column)

# 如何写入Excel
# 实例化一个工作簿对象
wb = Workbook()
# 找到当前活动的sheet
sheet = wb.active
# 修改一个title
sheet.title = "Shit"
# 填入内容
sheet["A3"] = "Hello World!"
# 循环填入
for i in range(10):
    sheet["B{}".format(i+1)] = i+1
# 利用Excel中的公式
sheet["A5"] = "=SUM(B:B)"
# 最后保存
wb.save("c:\\xxx.xlsx")

