# 此模块提供excel操作相关的功能
from openpyxl import load_workbook

def read_test_case_data(file_path):
    '''
    此方法提供读取测试用例数据的功能
    '''
    # 定义总的参数列表、总数据列表
    param_list = []
    total_list = []
    # 打开数据文件，得到Workbook对象
    workbook = load_workbook(file_path)
#     print(file_path)
    # 得到当前活动页
#     worksheet = workbook.get_active_sheet()
    worksheet = workbook.active
    # 得到最大列，以此确定表的标题内容
    columns = worksheet.max_column
    # 循环取出列的值，并存在列表中
    for i in range(0,columns):
        # 取出标题
        title = worksheet["1"][i].value
        # 以列表方式存储所有表头
        param_list.append(title)
    # 接下来得到最大行数
    rows = worksheet.max_row
    # 循环给取出所有数据并存在列表中
    for row in range(2,rows+1):
        # 定义单行数据列表，用于存储单行数据
        data_list = []
        for col in range(0,columns):
            cell_value = worksheet[row][col].value
            # 将空值的None变为空字符串
            if cell_value is None:
                cell_value=''
            data_list.append(cell_value)
        # 将单行数据加入总数据列表中
        total_list.append(data_list)
    # 定义字典存储参数列表和总数据列表
    data_dict = {}
    data_dict["param_list"] = param_list
    data_dict["total_list"] = total_list
#     print(data_dict)
    # 返回总数据列表
    return data_dict


if __name__ == "__main__":
    print(read_test_case_data("..\data\TestLogin.xlsx"))